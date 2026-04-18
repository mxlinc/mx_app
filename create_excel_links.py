import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Define the target directory
target_dir = r"C:\Users\abbas\ME Inc EXT\0000-MONTESSORI ONLINE\00-AllLinks"

# Create a workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Get all subdirectories
subdirs = []
try:
    for item in os.listdir(target_dir):
        item_path = os.path.join(target_dir, item)
        if os.path.isdir(item_path):
            subdirs.append(item)
except Exception as e:
    print(f"Error accessing directory: {e}")
    exit(1)

if not subdirs:
    print("No subdirectories found in the target directory.")
    exit(1)

# Sort subdirectories alphabetically
subdirs.sort()

# Process each subdirectory
for subdir in subdirs:
    subdir_path = os.path.join(target_dir, subdir)
    
    try:
        # Get only .url files in the subdirectory (non-recursive)
        files = []
        for item in os.listdir(subdir_path):
            item_path = os.path.join(subdir_path, item)
            if os.path.isfile(item_path) and item.lower().endswith('.url'):
                files.append(item)
        
        # Sort files alphabetically
        files.sort()
        
        # Create sheet name with subfolder name and file count
        file_count = len(files)
        sheet_name = f"{subdir} <{file_count}>"
        
        # Excel sheet names have a 31 character limit
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "...>"
        
        # Create a new sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add header
        header_cell = ws['A1']
        header_cell.value = "Filename"
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Add files to the sheet
        for idx, filename in enumerate(files, start=2):
            ws[f'A{idx}'].value = filename
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 50
        
        print(f"Sheet created: {sheet_name} with {file_count} files")
    
    except Exception as e:
        print(f"Error processing subdirectory '{subdir}': {e}")

# Save the workbook
output_path = os.path.join(target_dir, "All links.xlsx")
try:
    wb.save(output_path)
    print(f"\nExcel file saved successfully: {output_path}")
except Exception as e:
    print(f"Error saving Excel file: {e}")
